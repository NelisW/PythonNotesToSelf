"""

This script reads an Excel config file and one or more CSV or GTV matlab 
files and then proceeds to create and serve a Dash portal. 
The page served has several elements, all constructed from the 
information provided in the config file.

The config file has any number of sheets where each sheet defines
a different line graph (except for the header sheet, which defines 
the page header.)
Each graph sheet defines the title and height of the graph, axes labels,
one x-value column name and any number of y-value column names.
Each line has a number of attributes with default values if not supplied.

The data file is read and a set of Dash data structures are formed
according to the Excel config file specifications.

In the present script the config filename is hard coded as './pyqt-dash-config.xlsx'.
Adapt this script according to your needs.

Dash starts a Flask server at the specified port, so the browser must be 
pointing to the appropriate port number
localhost:port
This means that once the server is running, you can view the page with 
the PyQt browser as used here, or in an external browser.

There are numerous Dash and Plotly resources on the Internet:
https://dash.plot.ly/integrating-dash
https://www.datacamp.com/community/tutorials/learn-build-dash-python
https://github.com/plotly/dash-recipes
https://github.com/plotly/dash-recipes/blob/master/multiple-hover-data.py
https://plot.ly/python/subplots/

This script requires openpyxl, PyQt5, numpy, scipy, pandas, plotly and dash modules.

To install dash when connected to the internet:
conda config --add channels conda-forge
conda search dash-daq --channel conda-forge
conda install dash
conda install dash-html-components
conda install dash-core-components
conda install dash-table
conda install dash-daq

A recent off-line install required the following packages to be manually installed.
conda install dash-0.39.0-py_0.tar.bz2
conda install flask-compress-1.4.0-py_0.tar.bz2
conda install plotly-4.1.1-py_0.tar.bz2
conda install dash-html-components-0.14.0-py_0.tar.bz2
conda install dash-core-components-0.44.0-py_0.tar.bz2
conda install dash-table-3.6.0-py_0.tar.bz2
conda install dash-daq-0.1.4-py_0.tar.bz2
conda install plotly-orca-1.2.1-1.tar.bz2
conda install retrying-1.3.3-py36_1.tar.bz2
conda install dash-renderer-0.20.0-py_0.tar.bz2

Plotly packages seem to be here:  
https://anaconda.org/plotly  
https://anaconda.org/plotly/repo  
There are 17 packages, located under the package name, Files tab:
https://anaconda.org/plotly/plotly/files  
or   
https://anaconda.org/plotly/dash/files 

"""

import sys, os
import threading
import pandas as pd
import openpyxl as oxl
import numpy as np

from PyQt5 import QtWidgets
import PyQt5.QtCore as QtCore
from PyQt5 import QtWebEngineWidgets

import dash
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output, State

external_stylesheets = ['assets/bWLwgP.css']

pd.set_option('display.max_rows', 500)

# nice links:
# https://towardsdatascience.com/creating-an-interactive-data-app-using-plotlys-dash-356428b4699c
# https://dash.plot.ly/dash-core-components/tabs
# https://dash.plot.ly/getting-started-part-2

##########################################
#
def makeGraphSet(dft, graph):

    # get the header info from header sheet
    pagetitle = dfc.loc['Pagetitle','Value'] if 'Pagetitle' in dfc.index else ''
    pagetop = dfc.loc['PageTop','Value'] if 'PageTop' in dfc.index else ''
    pagebottom = dfc.loc['PageBottom','Value'] if 'PageBottom' in dfc.index else ''

    # list of all the line entries for this graph set
    graphData = []
    # dictionary to link lines and yValues
    yValueDict = {}
    numLines = 0
    # build the data for all lines in this graph set
    for index,row in dft[(dft['Variable']=='yValue')].iterrows():

        # add to yValue line dictionary
        yValueDict[numLines] = str(index)
        numLines = numLines + 1

        # get the filename for this graph
        dfilename = dft[(dft['Variable']=='Datafile')]['Value'].values[0]
        # get dataframe for this graph
        df = datafiles[dfilename]

        # y scale
        if 'Scale' in row:
            if not np.isnan(row['Scale']):
                yscale = row['Scale']
            else:
                yscale = 1.0
        # x scale
        if not np.isnan(dft[(dft['Variable']=='xValue')]['Scale'][0]):
            xscale = float(dft[(dft['Variable']=='xValue')]['Scale'][0])
        else:
            xscale = 1.0

        # y offset
        if 'Offset' in row:
            if not np.isnan(row['Offset']):
                yoffset = row['Offset']
            else:
                yoffset = 0.
        # x offset
        if not np.isnan(dft[(dft['Variable']=='xValue')]['Offset'][0]):
            xoffset = float(dft[(dft['Variable']=='xValue')]['Offset'][0])
        else:
            xoffset = 0.
                
        # each line in each graph must be a dict as follows:
        dLines = {
            'x':df[dft[(dft['Variable']=='xValue')].loc['xValue','Value']] * xscale + xoffset,
            'y':df[row['Value']] * yscale + yoffset,
            'line':{}
        }

        # fill in non-default values
        if not np.isnan(row['Linewidth']):
            dLines['line']['width'] = row['Linewidth']

        if isinstance(row['Colour'], str):
            dLines['line']['color'] = row['Colour']

        if isinstance(row['Dash'], str):
            dLines['line']['dash'] = row['Dash']

        if isinstance(row['GraphType'], str):
            dLines['type'] = row['GraphType']

        if isinstance(row['LineLabel'], str):
            dLines['name'] = row['LineLabel']
        else:
            dLines['name'] = row['Value']

        dLines['showlegend'] = True

        # add this line to other lines in this graph
        graphData.append(dLines)

    thisGraphList = []

    # append the text at the top of the graph
    thisGraphList.append(
        html.Div([dcc.Markdown(children=pagetop)]),
    )  
    if 'GraphTop' in dft.index:
        thisGraphList.append(
            html.Div([dcc.Markdown(children=dft.loc['GraphTop','Value'])])
        )            

    # run through all graphs in this set
    for index,row in dft[(dft['Variable']=='Title')].iterrows():
        # get the set number
        setStr = str(index).split('#')[1]
        setNumber = int(setStr)
        yValueName = 'yValue#'+setStr
        yLabelName = 'yLabel#'+setStr


        thisGraphData = []
        for (key, value) in yValueDict.items():
            if yValueName in value:
                thisGraphData.append(graphData[int(key)])


        figdict = {'layout':{'title': row['Value'],
                                'xaxis':{'title':dft.loc['xLabel','Value']},
                                'yaxis':{'title':dft.loc[yLabelName,'Value']},
                                },
                    'data':thisGraphData}

        #----------------------------------------------------------------------------------------------
        # appends the graph blocks to the list of graphs (top, graph, bottom)
        
        # append the actual graph
        thisGraphList.append(
            html.Div([dcc.Graph(
            id=graph+setStr,
            figure=figdict,
            style={'height': str(dft.loc['Height','Value'])},
            )])
        )

        # graphs to disk
        toDisk = True
        if 'ToDisk' in dft.index:
            if not np.isnan(dft[(dft['Variable']=='ToDisk')]['Value'].values[0]):
                toDisk = dft[(dft['Variable']=='ToDisk')]['Value'].values[0]
        
        if toDisk:
            # Save the figure to disk as html
            import plotly.offline as offline
            offline.plot(figdict,
                auto_open=False, 
                output_type='file', filename=f'{graph}#{setStr}.html', validate=False)

    # append the text at the bottom of the graph
    if 'GraphBottom' in dft.index:
        thisGraphList.append(
            html.Div([dcc.Markdown(children=dft.loc['GraphBottom','Value'])])
        )
    thisGraphList.append(
        html.Div([dcc.Markdown(children=pagebottom)]),
    )  
    return thisGraphList


##########################################
#
def prepareGraphs():

    # get list of all graphs in dfg
    graphs = dfg['Graph'].unique()

    # graphSets to be used when constructing the page
    # each entry in this list is a different tab containing several graphs
    global graphSets 
    global graphTabs
    graphSets = []
    graphTabs = []

    for graph in graphs:

        # extract info for this graph set
        dft = dfg[(dfg['Graph']==graph)]        
        
        # First check exclude flag
        toInclude = True
        if 'Include' in dft.index:
            if not np.isnan(dft[(dft['Variable']=='Include')]['Value'].values[0]):
                toInclude = dft[(dft['Variable']=='Include')]['Value'].values[0]

        if toInclude:
            graphSets.append(makeGraphSet(dft, graph))
            graphTabs.append(graph.split('-')[1])


##########################################
#
def makePage():
    """
    Create the contents to be displayed in the browser
    """

    # lstgraphs to be used when constructing the page
    # each entry in this list is a different tab containing several graphs
    lstgraphs = []

    for tabNum, tabSet in enumerate(graphSets):

        # --------------- now add to the tab
        # 
        tabLabel = graphTabs[tabNum]

        if useCallbacks:
            lstgraphs.append(
                dcc.Tab(value='Tab ' + str(tabNum), label=tabLabel))
        else:
            lstgraphs.append(
                dcc.Tab(value='Tab ' + str(tabNum), label=tabLabel, children=[
                    *tabSet,
                ]))


    # create the page to be rendered in the browser
    page = html.Div([
        dcc.Tabs(id="tabs", value='Tab 0', children=[
            # following is a list of dcc.Graph(() graphs
            *lstgraphs,
        ]),
        html.Div(id='tabs-content'),
    ])
    return page

##########################################
#
def run_dash(pageLayout,port):
    """Initiate the Dash server and serve the page

    pageLayout: info the be served in Plotly data format
    port: port number to be used
    """
    # start a dash app, which also starts a Flask server
    # global app
    app = dash.Dash(__name__, external_stylesheets=external_stylesheets)
    # override security restrictions: allow the serving of local pages
    app.css.config.serve_locally = True
    app.scripts.config.serve_locally = True
    app.layout = pageLayout

    @app.callback(Output('tabs-content', 'children'),
                [Input('tabs', 'value')])
    # def update_output(tab):
    def render_content(tab):
        tabNum = int(tab.split(' ')[1])
        return graphSets[tabNum]

    app.run_server(debug=False, port=port)

##########################################
#
# creates a browser widget
class WebViewer(QtWebEngineWidgets.QWebEngineView):
    def __init__(self, parent, url):
        """Create a web browser widget

        parent: the parent GUI element where this is included.
        url: the url to be browsed
        """
        super().__init__(parent)
        page = QtWebEngineWidgets.QWebEnginePage(self)
        self.setPage(page)
        self.setUrl(QtCore.QUrl(url))

##########################################
#
def loadConfig(configfile):
    """
    Loads the graph configuration from the excel file
        
    configfile: Excel file that defines the plots

    """

    # read the config file
    cxls = pd.ExcelFile(configfile)

    # header dataframe
    global dfc
    dfc = pd.read_excel(cxls, 'header')
    dfc = dfc.set_index('Variable')

    # get a list of graph sheetnames (ignore the header sheet)
    cwb =  oxl.load_workbook(configfile)
    sheetnames = [sn for sn in cwb.sheetnames if 'graph' in sn]

    # global dataframe to contain ALL the sheets' info
    global dfg
    dfg = pd.DataFrame()

    for shtnum,sheetname in enumerate(sheetnames):
        dft = pd.read_excel(cxls, sheetname)
        # add info to identify the lines associated with this sheet
        dft['Graph'] = sheetname
        dft['ShtNum'] = shtnum
        dft['Index'] = dft['Variable']

        # determine the number of graphs on this tab
        i = 0
        theSet = -1
        for index,row in dft.iterrows():
            if 'Title' in row['Variable']:
                theSet = theSet + 1
                dft.loc[index,'Index'] = f"{row['Variable']}#{theSet:03d}"
            if 'yLabel' in row['Variable']:
                dft.loc[index,'Index'] = f"{row['Variable']}#{theSet:03d}"
                i = 0
            if 'yValue' in row['Variable']:
                dft.loc[index,'Index'] = f"{row['Variable']}#{theSet:03d}-{i:03d}"
                i = i + 1

        # make 'Index' column the index
        dft = dft.set_index('Index')
        # append this sheet to the master data frame
        dfg = dfg.append(dft)

    # print(dfg)

##########################################
#
def loadData():
    """
    Load all the telemetry data from all files required
    """

    # get data filenames in all the graphs
    datafilenames = dfg[(dfg['Variable']=='Datafile')]['Value'].unique()

    global datafiles
    datafiles = {}

    for datafilename in datafilenames:
        extension = os.path.splitext(datafilename)[1]

        if 'mat' in extension:

            # load the gtv telemetry data in matlab format file
            # scipy reads in structures as structured numpy arrays of dtype object
            from scipy.io import loadmat
            dataMat = loadmat(datafilename)

            # create the dataframe
            datafiles[datafilename] = pd.DataFrame(dataMat['DATA'], columns=dataMat['NAM'])

            # set begin of ATP as time zero
            counter = datafiles[datafilename]['host_bfAtpEvents'][datafiles[datafilename]['host_bfAtpEvents'] == 1].index.values.astype(int)[0]
            datafiles[datafilename]['TIME'] = datafiles[datafilename]['TIME'] - datafiles[datafilename]['TIME'][counter]
            
        else:
            # read the data file
            datafiles[datafilename] = pd.read_csv(datafilename, sep="\s+|,|;", index_col=None,engine='python')
 
##########################################
#
if __name__ == "__main__":
       
    try:
        from docopt import docopt
    except ImportError:
        print('Install docopt using Anaconda:')
        print('    conda install -c anaconda docopt')
        print('or if not using Anaconda: ')
        print('    pip install docopt')
        print('or simply put the docopt.py script in the working folder')
        sys.exit(0)

    options = """pyqt-dash-lineplot.py

        Usage:
          pyqt-dash-lineplot.py [--configfile=<configFilename>] [--nocallback] [-n] 
          pyqt-dash-lineplot.py -h | --help 
          pyqt-dash-lineplot.py -n | --nocallback
 
        Options:
          -h, --help                           Show this screen.
          -n, --nocallback                     Run app with callbacks or not default: False].         
          -f <configFilename>, --configfile <configFilename>    Excel config filename [default: ./pyqt-dash-config.xlsx].
 
    """
    # process commandline arguments
    optionArguments = docopt(options)

    # make callbacks flag global
    global useCallbacks
    useCallbacks = not optionArguments["--nocallback"]

    sys.argv.append("--disable-web-security")
    
    port = '8050' # used for the local Flask server

    # Excel file that defines the plots
    configfile = optionArguments["--configfile"]
    loadConfig(configfile)

    # load all data to be available globally 
    # all the data files, but only once into a dict with filename as key
    loadData()
    
    # prepare all required graph sets
    prepareGraphs()
  
    # now create the page we want to render
    pageLayout = makePage() 

    # start the thread
    threading.Thread(target=run_dash, args=(pageLayout,port), daemon=True).start()

    # start main app with the widget rendering what is available on the port
    appMain = QtWidgets.QApplication(sys.argv)
    main_widget = QtWidgets.QWidget(None)
    window_layout = QtWidgets.QVBoxLayout(main_widget)
    window_layout.addWidget(WebViewer(main_widget,f'http://127.0.0.1:{port}'))
    main_widget.show()

    sys.exit(appMain.exec_())
