"""

This script reads a CSV file with a header and an Excel config file 
that defines the dash page layout, plotting a number of line graphs.

The config file has any number of sheets where each sheet defines
a different line graph (except for the header sheet, which defines 
the page header.)
Each graph sheet defines the title and height of the graph, 
one x-value column name and any number of y-value column names.

The data file is read and a set of Dash data structures are formed
according to the Excel config file specifications.

In the present script the config filename and data filename are hard coded.
Adapt this script according to your needs.


https://github.com/plotly/dash-recipes
https://github.com/plotly/dash-recipes/blob/master/multiple-hover-data.py
https://plot.ly/python/subplots/

conda config --add channels conda-forge
conda search dash-daq --channel conda-forge

conda install dash
conda install dash-html-components
conda install dash-core-components
conda install dash-table
conda install dash-daq

Dash starts a Flask server at the specified port, so the browser must be 
pointing to the appropriate port number
localhost:port
This means that once the server is running, you can view the page with 
the PyQt browser as used here, or in an external browser.
https://dash.plot.ly/integrating-dash
https://www.datacamp.com/community/tutorials/learn-build-dash-python

"""

import sys
import threading
import pandas as pd
import openpyxl as oxl

from PyQt5 import QtWidgets
import PyQt5.QtCore as QtCore
from PyQt5 import QtWebEngineWidgets

import dash
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output
# import plotly.graph_objs as go


# creates a browser widget
class WebViewer(QtWebEngineWidgets.QWebEngineView):
    def __init__(self, parent, url):
        """Create a web browser widget

        parent: the parent GUI element where this is included.
        url: the url to be browsed
        """
        super().__init__(parent)
        page = QtWebEngineWidgets.QWebEnginePage(self)
        self.setPage(page)
        self.setUrl(QtCore.QUrl(url))


def makeDccGraph(gid, height, title, data):
    """Builds the data structure for one set of line graphs.

    id: must be unique to differentiate between graphs.
    figure: defines the graph layout in terms of
        title 
        data set (a complex data structure)
    style: defines some properties of the graph.

    This function can probably be expanded for improved style

    The function returns a dcc.Graph()function with parameters.
    """

    return dcc.Graph(
                id=gid,
                figure={'layout':{'title':title},'data':data},
                style={'height': str(height)},
            )


def makeHTMLPage(pagetitle,pageheader,lstgraphs):
    """Builds the complete page layout to be used in the browser.

    pagetitle: the text displayed in H1 format at the top of page.
    pageheader: a Markdown text block displayed immediately after 
                the title.
    lstgraphs: a list of data structures for the different graphs
               on the page. The list of structures is unpacked when 
               html.Div is called.  This means that what is passed 
               as a list is unpacked into separate function arguments
               by the *list operator.
    """
    # build a page for display
    page = html.Div(
        children=[
            html.H1(children=pagetitle),
            html.Div([dcc.Markdown(children=pageheader)]),
            # following is a list of dcc.Graph(() graphs
            *lstgraphs
        ]
    )
    return page



#####################################################
def makePageFromCVS(configfile):
    """Create the contents to be displayed in the browser

    configfile: the Excel file that defines the plots

    """

    #read the config file
    cxls = pd.ExcelFile(configfile)
    dfc = pd.read_excel(cxls, 'header')
    dfc = dfc.set_index('Variable')

    # get a list of graph sheetnames (ignore the header sheet)
    cwb =  oxl.load_workbook(configfile)
    sheetnames = [sn for sn in cwb.get_sheet_names() if 'graph' in sn]
    # dataframe to contain ALL the sheets' info
    dfg = pd.DataFrame()

    # build dfg with all sheets' info
    for shtnum,sheetname in enumerate(sheetnames):
        dft = pd.read_excel(cxls, sheetname)
        # add info to identify the lines associated with this sheet
        dft['Graph'] = sheetname
        dft['ShtNum'] = shtnum
        # create unique index values by adding number of the yValues string
        dft['Index'] = dft['Variable']
        i = 0
        for index,row in dft.iterrows():
            print(row['Variable'])
            if 'yValue' in row['Variable']:
                dft.loc[index,'Index'] = f"{row['Variable']}{i:03d}"
                i = i + 1
        # make 'Index' column the index
        dft = dft.set_index('Index')
        # append this sheet to the master data frame
        dfg = dfg.append(dft)

    print(dfg)

    # get data filenames in all the graphs
    datafilenames = dfg[(dfg['Variable']=='Datafile')]['Value'].unique()
    print(datafilenames)
    # load all the data files, but only once into a dict with filename as key
    datafiles = {}
    for datafilename in datafilenames:
        # read the data file
        datafiles[datafilename] = pd.read_csv(datafilename, sep="\s+|,|;", index_col=None,engine='python')

    # now we have one dataframe with the columns:
    # Variable, Value , Name, LineType, Graph, ShtNum

    # get list of all graphs in dfg
    graphs = dfg['Graph'].unique()
    # lstgraphs to be used when constructing the page
    # each entry in this list is a different graph
    lstgraphs = []
    for graph in graphs:
        # extract info for this graph
        dft = dfg[(dfg['Graph']==graph)]

        # get name for the data used as x-values
        dfx = dft[(dft['Variable']=='xValue')]
        # for all yValue lines in this graph
        dfy = dft[(dft['Variable']=='yValue')]
        # list of all the line entries for this graph
        grpData = []
        # build the data for all lines in this graph
        for index,row in dfy.iterrows():
            # get the filename for this graph
            dfilename = dft[(dft['Variable']=='Datafile')]['Value'].values[0]
            # get dataframe for this graph
            df = datafiles[dfilename]

            # each line in each graph must be a dict as follows:
            grpData.append({
                # 'x':df[dfx.iloc[0]['Value']],
                'x':dft.loc['xValue','Value'], #/ float(row['Scale']),
                'y':df[row['Value']], #/ float(row['Scale']),
                'type':row['LineType'],
                'name':row['Name'],
            })

        # create the plotly data structure for this graph (include title.height)
        lstgraphs.append(makeDccGraph(gid=graph, 
                                    height=dft.loc['Height','Value'], 
                                    title=dft.loc['Title','Value'], 
                                    data=grpData))

    # get the header info from header sheet
    pageheader = dfc.loc['Pageheader','Value']
    pagetitle = dfc.loc['Pagetitle','Value']

    # create the page to be renderer in the browser
    page = makeHTMLPage(pagetitle,pageheader,lstgraphs)

    return page


def run_dash(pageLayout,port):
    """Initiate the Dash server and serve the page

    pageLayout: info the be served in Plotly data formal
    port: port number to be used
    """
    # start a dash app, which also starts a Flask server
    app = dash.Dash()
    # override security restrictions: allow the serving of local pages
    app.css.config.serve_locally = True
    app.scripts.config.serve_locally = True
    app.layout = pageLayout
    # app.css.append_css({
    #     'external_url': 'https://codepen.io/chriddyp/pen/bWLwgP.css'
    # })
    app.run_server(debug=False, port=port)


##########################################
#

if __name__ == "__main__":

    sys.argv.append("--disable-web-security")
    
    port = '8050' # used for the local Flask server

    configfile = 'data/pyqt-dash-config.xlsx'
    pageLayout = makePageFromCVS(configfile) 
    threading.Thread(target=run_dash, args=(pageLayout,port), daemon=True).start()

    app = QtWidgets.QApplication(sys.argv)
    main_widget = QtWidgets.QWidget(None)
    window_layout = QtWidgets.QVBoxLayout(main_widget)
    window_layout.addWidget(WebViewer(main_widget,f'http://127.0.0.1:{port}'))
    main_widget.show()

    sys.exit(app.exec_())
