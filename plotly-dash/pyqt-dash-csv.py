"""

This script reads an Excel config file and one or more CSV files and 
then proceeds to create and serve a Dash portal. The page served has
several elements, all constructed from the information provided in the 
config file.

The config file has any number of sheets where each sheet defines
a different line graph (except for the header sheet, which defines 
the page header.)
Each graph sheet defines the title and height of the graph, axes labels,
one x-value column name and any number of y-value column names.
Each line has a number of attributes with default values if not supplied.

The data file is read and a set of Dash data structures are formed
according to the Excel config file specifications.

In the present script the config filename and data filename are hard coded.
Adapt this script according to your needs.

See the example config file in the repository at data/pyqt-dash-config.xlsx.

Dash starts a Flask server at the specified port, so the browser must be 
pointing to the appropriate port number
localhost:port
This means that once the server is running, you can view the page with 
the PyQt browser as used here, or in an external browser.

There are numerous Dash and Plotly resources on the Internet:
https://dash.plot.ly/integrating-dash
https://www.datacamp.com/community/tutorials/learn-build-dash-python
https://github.com/plotly/dash-recipes
https://github.com/plotly/dash-recipes/blob/master/multiple-hover-data.py
https://plot.ly/python/subplots/

This script requires openpyxl, PyQt5, numpy, pandas, plotly and dash modules.

To install dash when connected to the internet:
conda config --add channels conda-forge
conda search dash-daq --channel conda-forge
conda install dash
conda install dash-html-components
conda install dash-core-components
conda install dash-table
conda install dash-daq

"""

import sys
import threading
import pandas as pd
import openpyxl as oxl
import numpy as np

from PyQt5 import QtWidgets
import PyQt5.QtCore as QtCore
from PyQt5 import QtWebEngineWidgets

import dash
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output
# import plotly.graph_objs as go


#####################################################
def makePageFromCVS(configfile):
    """Create the contents to be displayed in the browser

    configfile: the Excel file that defines the plots
    """

    #read the config file
    cxls = pd.ExcelFile(configfile)
    dfc = pd.read_excel(cxls, 'header')
    dfc = dfc.set_index('Variable')

    # get a list of graph sheetnames (ignore the header sheet)
    cwb =  oxl.load_workbook(configfile)
    # sheetnames = [sn for sn in cwb.get_sheet_names() if 'graph' in sn]
    sheetnames = [sn for sn in cwb.sheetnames if 'graph' in sn]
    # dataframe to contain ALL the sheets' info
    dfg = pd.DataFrame()

    # build dfg with all sheets' info
    for shtnum,sheetname in enumerate(sheetnames):
        dft = pd.read_excel(cxls, sheetname)
        # add info to identify the lines associated with this sheet
        dft['Graph'] = sheetname
        dft['ShtNum'] = shtnum
        # create unique index values by adding number of the yValues string
        dft['Index'] = dft['Variable']
        i = 0
        for index,row in dft.iterrows():
            if 'yValue' in row['Variable']:
                dft.loc[index,'Index'] = f"{row['Variable']}{i:03d}"
                i = i + 1
        # make 'Index' column the index
        dft = dft.set_index('Index')
        # append this sheet to the master data frame
        dfg = dfg.append(dft)

    # print(dfg)

    # get data filenames in all the graphs
    datafilenames = dfg[(dfg['Variable']=='Datafile')]['Value'].unique()

    # load all the data files, but only once into a dict with filename as key
    datafiles = {}
    for datafilename in datafilenames:
        # read the data file
        datafiles[datafilename] = pd.read_csv(datafilename, sep="\s+|,|;", index_col=None,engine='python')

    # get list of all graphs in dfg
    graphs = dfg['Graph'].unique()
    # lstgraphs to be used when constructing the page
    # each entry in this list is a different graph
    lstgraphs = []
    for graph in graphs:
        # extract info for this graph
        dft = dfg[(dfg['Graph']==graph)]
        # list of all the line entries for this graph
        graphData = []
        # build the data for all lines in this graph
        for index,row in dft[(dft['Variable']=='yValue')].iterrows():
            # get the filename for this graph
            dfilename = dft[(dft['Variable']=='Datafile')]['Value'].values[0]
            # get dataframe for this graph
            df = datafiles[dfilename]

            # yscale
            if 'Scale' in row:
                if not np.isnan(row['Scale']):
                    yscale = row['Scale']
                else:
                    yscale = 1.0
            # x scale
            if not np.isnan(dft[(dft['Variable']=='xValue')]['Scale'][0]):
                xscale = float(dft[(dft['Variable']=='xValue')]['Scale'][0])
            else:
                xscale = 1.0
                    
            # each line in each graph must be a dict as follows:
            dLines = {
                'x':df[dft[(dft['Variable']=='xValue')].loc['xValue','Value']] * xscale,
                'y':df[row['Value']] * yscale,
                'line':{}
            }

            # fill in non-default values
            if not np.isnan(row['Linewidth']):
                dLines['line']['width'] = row['Linewidth']

            if isinstance(row['Colour'], str):
                dLines['line']['color'] = row['Colour']

            if isinstance(row['Dash'], str):
                dLines['line']['dash'] = row['Dash']

            if isinstance(row['GraphType'], str):
                dLines['type'] = row['GraphType']

            if isinstance(row['LineLabel'], str):
                dLines['name'] = row['LineLabel']
            else:
                dLines['name'] = row['Value']

            # add this line to other lines in this graph
            graphData.append(dLines)

        # appends the graph blocks to the list of graphs (top, graph, bottom)
        # append the text at the top of the graph
        if 'GraphTop' in dft.index:
            lstgraphs.append(html.Div([dcc.Markdown(children=dft.loc['GraphTop','Value'])]))
        # append the actual graph
        lstgraphs.append(html.Div([dcc.Graph(
                id=graph,
                figure={'layout':{'title':dft.loc['Title','Value'],
                                'xaxis':{'title':dft.loc['xLabel','Value']},
                                'yaxis':{'title':dft.loc['yLabel','Value']},
                                },
                    'data':graphData},
                style={'height': str(dft.loc['Height','Value'])},
            )]))
        # append the text at the bottom of the graph
        if 'GraphBottom' in dft.index:
            lstgraphs.append(html.Div([dcc.Markdown(children=dft.loc['GraphBottom','Value'])]))

    # get the header info from header sheet
    pagetitle = dfc.loc['Pagetitle','Value'] if 'Pagetitle' in dfc.index else ''
    pagetop = dfc.loc['PageTop','Value'] if 'PageTop' in dfc.index else ''
    pagebottom = dfc.loc['PageBottom','Value'] if 'PageBottom' in dfc.index else ''

    # create the page to be rendered in the browser
    page = html.Div(
        children=[
            html.H1(children=pagetitle),
            html.Div([dcc.Markdown(children=pagetop)]),
            # following is a list of dcc.Graph(() graphs
            *lstgraphs,
            html.Div([dcc.Markdown(children=pagebottom)]),
        ]
    )
    return page


def run_dash(pageLayout,port):
    """Initiate the Dash server and serve the page

    pageLayout: info the be served in Plotly data formal
    port: port number to be used
    """
    # start a dash app, which also starts a Flask server
    app = dash.Dash()
    # override security restrictions: allow the serving of local pages
    app.css.config.serve_locally = True
    app.scripts.config.serve_locally = True
    app.layout = pageLayout
    # app.css.append_css({
    #     'external_url': 'https://codepen.io/chriddyp/pen/bWLwgP.css'
    # })
    app.run_server(debug=False, port=port)


# creates a browser widget
class WebViewer(QtWebEngineWidgets.QWebEngineView):
    def __init__(self, parent, url):
        """Create a web browser widget

        parent: the parent GUI element where this is included.
        url: the url to be browsed
        """
        super().__init__(parent)
        page = QtWebEngineWidgets.QWebEnginePage(self)
        self.setPage(page)
        self.setUrl(QtCore.QUrl(url))

##########################################
#

if __name__ == "__main__":

    sys.argv.append("--disable-web-security")
    
    port = '8050' # used for the local Flask server

    configfile = 'data/pyqt-dash-config.xlsx'
    pageLayout = makePageFromCVS(configfile) 
    threading.Thread(target=run_dash, args=(pageLayout,port), daemon=True).start()

    app = QtWidgets.QApplication(sys.argv)
    main_widget = QtWidgets.QWidget(None)
    window_layout = QtWidgets.QVBoxLayout(main_widget)
    window_layout.addWidget(WebViewer(main_widget,f'http://127.0.0.1:{port}'))
    main_widget.show()

    sys.exit(app.exec_())
